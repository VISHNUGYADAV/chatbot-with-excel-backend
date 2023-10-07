import openpyxl
from chatterbot import ChatBot
from chatterbot.trainers import ChatterBotCorpusTrainer

def load_workbook(file_path):
    try:
        workbook = openpyxl.load_workbook(file_path)
        return workbook
    except Exception as e:
        return None

def get_sheet_names(workbook):
    return workbook.sheetnames

def get_sheet_data(workbook, sheet_name):
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data

def main():
    print("Excel Automation Chatbot")
    file_path = input("Enter the path to the Excel file: ")
    
    workbook = load_workbook(file_path)
    if workbook:
        chatbot = ChatBot('ExcelBot')
        trainer = ChatterBotCorpusTrainer(chatbot)
        trainer.train("chatterbot.corpus.english")
            
        print("Chat with the ExcelBot. Type 'exit' to end the conversation.")
        while True:
            user_input = input("You: ")
            if user_input.lower() == 'exit':
                break
            response = chatbot.get_response(user_input)
            print("ExcelBot:", response)
                
            if "list sheets" in user_input:
                sheet_names = get_sheet_names(workbook)
                response = f"Available sheets: {', '.join(sheet_names)}"
                print("ExcelBot:", response)
                
            if "read sheet" in user_input:
                sheet_names = get_sheet_names(workbook)
                response = f"Available sheets: {', '.join(sheet_names)}. Which sheet do you want to read?"
                print("ExcelBot:", response)
                user_input = input("You: ")
                sheet_name = user_input.strip()
                if sheet_name in sheet_names:
                    sheet_data = get_sheet_data(workbook, sheet_name)
                    response = f"Displaying data from sheet '{sheet_name}': {sheet_data}"
                    print("ExcelBot:", response)
                else:
                    response = "Invalid sheet name."
                    print("ExcelBot:", response)
                        
        else:
            print("Failed to load the Excel file.")

    if _name_ == "_main_":
    main()